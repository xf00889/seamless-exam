from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils import timezone
from django.urls import reverse
from django.core.paginator import Paginator
from itertools import groupby
from exams.models import Exam
from attempts.models import Attempt, AttemptStatus
from services.exam_service import ExamService
from services.attempt_service import AttemptService
from services.answer_service import AnswerService
from services.data_integrity_service import DataIntegrityService
from services.tab_monitoring_service import TabMonitoringService
from services.activity_log_service import ActivityLogService
from services.view_helpers import build_breadcrumbs
import json
import logging

logger = logging.getLogger('services')


# Initialize services
exam_service = ExamService()
attempt_service = AttemptService()
answer_service = AnswerService()
data_integrity_service = DataIntegrityService()
tab_monitoring_service = TabMonitoringService()
activity_log_service = ActivityLogService()


def _attempt_timestamp(attempt):
    return attempt.submitted_at or attempt.started_at


def _attempt_pending_essay_count(attempt):
    from exams.models import QuestionType

    pending_count = 0
    for answer in attempt.answers.all():
        if answer.question.question_type == QuestionType.ESSAY and answer.is_correct is None:
            pending_count += 1
    return pending_count


def _get_teacher_submission_attempts():
    return Attempt.objects.select_related(
        'student',
        'student__class_assigned',
        'exam',
        'exam__quarter',
    ).prefetch_related(
        'answers',
        'answers__question',
    ).filter(
        status__in=[AttemptStatus.SUBMITTED, AttemptStatus.GRADED]
    ).order_by(
        'student__last_name',
        'student__first_name',
        'student__school_id',
        '-submitted_at',
        '-started_at',
    )


def _build_student_rows(attempts):
    student_rows = {}

    for attempt in attempts:
        student = attempt.student
        latest_timestamp = _attempt_timestamp(attempt)
        pending_count = _attempt_pending_essay_count(attempt)

        if student.id not in student_rows:
            student_rows[student.id] = {
                'student': student,
                'student_name': student.get_full_name(),
                'student_class': student.class_assigned,
                'attempt_count': 0,
                'exam_ids': set(),
                'pending_attempt_count': 0,
                'latest_submission': latest_timestamp,
                'latest_attempt': attempt,
                'attempts': [],
            }

        row = student_rows[student.id]
        row['attempts'].append(attempt)
        row['attempt_count'] += 1
        row['exam_ids'].add(attempt.exam_id)
        row['pending_attempt_count'] += 1 if pending_count else 0

        if latest_timestamp and (row['latest_submission'] is None or latest_timestamp > row['latest_submission']):
            row['latest_submission'] = latest_timestamp
            row['latest_attempt'] = attempt

    rows = list(student_rows.values())
    for row in rows:
        row['exam_count'] = len(row['exam_ids'])
        del row['exam_ids']

    rows.sort(
        key=lambda row: (
            row['latest_submission'] or timezone.now(),
            row['student_name'].lower(),
        ),
        reverse=True,
    )
    return rows


def _build_exam_groups_for_student(attempts):
    exam_groups = {}

    for attempt in attempts:
        exam = attempt.exam
        latest_timestamp = _attempt_timestamp(attempt)
        pending_count = _attempt_pending_essay_count(attempt)

        if exam.id not in exam_groups:
            exam_groups[exam.id] = {
                'exam': exam,
                'attempt_count': 0,
                'pending_attempt_count': 0,
                'latest_submission': latest_timestamp,
                'attempts': [],
            }

        group = exam_groups[exam.id]
        group['attempts'].append({
            'attempt': attempt,
            'pending_essay_count': pending_count,
            'has_ungraded_essays': pending_count > 0,
            'latest_timestamp': latest_timestamp,
        })
        group['attempt_count'] += 1
        group['pending_attempt_count'] += 1 if pending_count else 0

        if latest_timestamp and (group['latest_submission'] is None or latest_timestamp > group['latest_submission']):
            group['latest_submission'] = latest_timestamp

    groups = list(exam_groups.values())
    for group in groups:
        group['attempts'].sort(
            key=lambda data: data['latest_timestamp'] or timezone.now(),
            reverse=True,
        )
        group['latest_attempt'] = group['attempts'][0]['attempt'] if group['attempts'] else None

    groups.sort(key=lambda group: group['exam'].title.lower())
    return groups


def _build_pending_grading_rows(attempts):
    pending_rows = []

    for attempt in attempts:
        pending_count = _attempt_pending_essay_count(attempt)
        if not pending_count:
            continue

        pending_rows.append({
            'attempt': attempt,
            'student': attempt.student,
            'exam': attempt.exam,
            'pending_essay_count': pending_count,
            'latest_timestamp': _attempt_timestamp(attempt),
        })

    pending_rows.sort(
        key=lambda row: (
            row['latest_timestamp'] or timezone.now(),
            row['student'].get_full_name().lower(),
        ),
        reverse=True,
    )
    return pending_rows


@ensure_csrf_cookie
def student_exam_list_view(request):
    """
    Display all active exams for students.
    Shows exam title, duration, question count, and completion status.
    Filters exams by student's assigned class.
    Requirements: 9.1, 10.1, 9.5, 3.2, 3.3
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        messages.error(request, 'Please log in to view exams')
        return redirect('student_login')
    
    student_id = request.session['student_id']
    
    # Get student to check class assignment (Requirement 3.2)
    from users.models import Student
    try:
        student = Student.objects.select_related('class_assigned').get(id=student_id)
    except Student.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('student_login')
    
    # Get all active exams with optimized queries (Requirement 9.5)
    # Prefetch questions to avoid N+1 queries when counting questions
    active_exams = exam_service.get_active_exams()
    
    # Filter exams by student's assigned class (Requirements 3.2, 3.3)
    if student.class_assigned:
        # Student has a class assignment - filter exams
        from exams.models import ExamClassAssignment
        
        # Get exams assigned to student's class
        assigned_exam_ids = set(
            ExamClassAssignment.objects.filter(
                class_assigned=student.class_assigned
            ).values_list('exam_id', flat=True)
        )
        
        # Also include exams with no class assignments (available to all)
        exams_with_assignments = set(
            ExamClassAssignment.objects.values_list('exam_id', flat=True).distinct()
        )
        
        # Filter: show exams assigned to student's class OR exams with no assignments
        filtered_exams = []
        for exam in active_exams:
            if exam.id in assigned_exam_ids or exam.id not in exams_with_assignments:
                filtered_exams.append(exam)
        
        active_exams = filtered_exams
    else:
        # Student has no class assignment - only show exams with no class restrictions (Requirement 3.3)
        from exams.models import ExamClassAssignment
        
        exams_with_assignments = set(
            ExamClassAssignment.objects.values_list('exam_id', flat=True).distinct()
        )
        
        # Only show exams that have no class assignments
        active_exams = [exam for exam in active_exams if exam.id not in exams_with_assignments]
    
    # Get student's attempts to mark completed exams
    student_attempts = attempt_service.get_student_attempts(student_id)
    completed_exam_ids = set()
    in_progress_attempts = {}
    graded_attempts = {}
    
    for attempt in student_attempts:
        if attempt.status == AttemptStatus.SUBMITTED or attempt.status == AttemptStatus.GRADED:
            completed_exam_ids.add(attempt.exam_id)
            if attempt.status == AttemptStatus.GRADED:
                graded_attempts[attempt.exam_id] = attempt
        elif attempt.status == AttemptStatus.IN_PROGRESS:
            in_progress_attempts[attempt.exam_id] = attempt
    
    # Check ExamStudentAssignment to filter exams by individual student access
    from repositories.exam_student_assignment_repository import ExamStudentAssignmentRepository
    from exams.models import ExamStudentAssignment
    
    assignment_repo = ExamStudentAssignmentRepository()
    
    # Get exams with student assignments
    active_exam_ids = [e.id for e in active_exams]
    exams_with_student_assignments = set(
        ExamStudentAssignment.objects.filter(exam_id__in=active_exam_ids)
        .values_list('exam_id', flat=True)
        .distinct()
    )
    
    # Prepare exam data with completion status
    exam_data = []
    for exam in active_exams:
        # Check if exam has student assignments
        if exam.id in exams_with_student_assignments:
            # Exam has individual student assignments - check if this student has access
            if not assignment_repo.is_student_assigned(exam.id, student_id):
                # Student doesn't have access, skip this exam
                continue
        
        # Student has access (either no assignments exist, or student is assigned)
        exam_info = {
            'exam': exam,
            'is_completed': exam.id in completed_exam_ids,
            'in_progress_attempt': in_progress_attempts.get(exam.id),
            'graded_attempt': graded_attempts.get(exam.id),
            'question_count': exam.questions.count()
        }
        exam_data.append(exam_info)

    exam_data.sort(key=lambda data: (
        data['exam'].quarter.order if data['exam'].quarter else 9999,
        data['exam'].quarter.name.lower() if data['exam'].quarter else 'no quarter',
        -data['exam'].created_at.timestamp(),
    ))

    grouped_quarters = []
    for key, group in groupby(
        exam_data,
        key=lambda data: (
            data['exam'].quarter.id if data['exam'].quarter else None,
            data['exam'].quarter.name if data['exam'].quarter else 'No Quarter',
            data['exam'].quarter.order if data['exam'].quarter else 9999,
        )
    ):
        quarter_items = list(group)
        grouped_quarters.append({
            'quarter_id': key[0],
            'label': key[1],
            'exam_count': len(quarter_items),
            'exams': quarter_items,
        })
    
    # Build breadcrumbs
    breadcrumbs = build_breadcrumbs(
        'Available Exams'
    )
    
    context = {
        'exam_data': exam_data,
        'grouped_quarters': grouped_quarters,
        'has_exams': len(exam_data) > 0,
        'student_class': student.class_assigned,
        'breadcrumbs': breadcrumbs
    }
    return render(request, 'attempts/student_exam_list.html', context)


@require_http_methods(["POST"])
def exam_start_view(request, exam_id):
    """
    Initialize an exam attempt for a student.
    Creates a new attempt record and starts the timer.
    Requirements: 10.1
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Check if exam exists and is active
    exam = exam_service.get_exam(exam_id)
    if not exam:
        return JsonResponse({'error': 'Exam not found'}, status=404)
    
    if not exam.is_active:
        return JsonResponse({'error': 'Exam is not active'}, status=400)
    
    # Check ExamStudentAssignment for individual student access
    from repositories.exam_student_assignment_repository import ExamStudentAssignmentRepository
    assignment_repo = ExamStudentAssignmentRepository()
    
    # If exam has student assignments, verify this student has access
    if assignment_repo.has_any_assignments(exam_id):
        if not assignment_repo.is_student_assigned(exam_id, student_id):
            return JsonResponse({'error': 'You do not have access to this exam'}, status=403)
    
    # Check if student already has a completed attempt
    existing_attempts = attempt_service.get_student_exam_attempts(student_id, exam_id)
    for attempt in existing_attempts:
        if attempt.status in [AttemptStatus.SUBMITTED, AttemptStatus.GRADED]:
            return JsonResponse({'error': 'You have already completed this exam'}, status=400)
    
    # Check if there's an in-progress attempt
    in_progress_attempt = attempt_service.get_in_progress_attempt(student_id, exam_id)
    if in_progress_attempt:
        # Resume existing attempt
        return JsonResponse({
            'success': True,
            'attempt_id': in_progress_attempt.id,
            'message': 'Resuming existing attempt'
        })
    
    # Create new attempt
    attempt = attempt_service.create_attempt(student_id, exam_id)
    
    if attempt:
        return JsonResponse({
            'success': True,
            'attempt_id': attempt.id,
            'message': 'Exam started successfully'
        })
    else:
        return JsonResponse({'error': 'Failed to start exam'}, status=500)


def exam_take_view(request, attempt_id):
    """
    Display exam questions and handle answer submission.
    Shows timer, questions, and allows answer input.
    Implements answer immutability after submission.
    Requirements: 10.1, 11.5
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        messages.error(request, 'Please log in to take exams')
        return redirect('student_login')
    
    student_id = request.session['student_id']
    
    # Get attempt with full data
    attempt = attempt_service.get_attempt_with_full_data(attempt_id)
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('student_exam_list')
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        messages.error(request, 'You do not have permission to access this attempt')
        return redirect('student_exam_list')
    
    # Check if attempt is already submitted (answer immutability)
    if attempt.status != AttemptStatus.IN_PROGRESS:
        messages.info(request, 'This exam has already been submitted')
        return redirect('student_exam_list')
    
    # Get exam with questions
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('student_exam_list')
    
    # Get existing answers
    existing_answers = answer_service.get_attempt_answers(attempt_id)
    answer_dict = {answer.question_id: answer for answer in existing_answers}
    
    # Calculate remaining time
    elapsed_time = (timezone.now() - attempt.started_at).total_seconds()
    total_time = exam.duration_minutes * 60
    remaining_time = max(0, total_time - elapsed_time)
    
    # Prepare questions with answers
    questions_data = []
    for question in exam.questions.all():
        question_info = {
            'question': question,
            'answer': answer_dict.get(question.id)
        }
        questions_data.append(question_info)
    
    context = {
        'attempt': attempt,
        'exam': exam,
        'questions_data': questions_data,
        'remaining_time': int(remaining_time),
        'duration_minutes': exam.duration_minutes,
        'total_questions': len(questions_data)
    }
    return render(request, 'attempts/exam_take.html', context)


def exam_review_view(request, attempt_id):
    """
    Display review page showing all questions and answers before final submission.
    Read-only view allowing students to review their answers.
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        messages.error(request, 'Please log in to view this page')
        return redirect('student_login')
    
    student_id = request.session['student_id']
    
    # Get attempt with full data
    attempt = attempt_service.get_attempt_with_full_data(attempt_id)
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('student_exam_list')
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        messages.error(request, 'You do not have permission to access this attempt')
        return redirect('student_exam_list')
    
    # Check if attempt is already submitted
    if attempt.status != AttemptStatus.IN_PROGRESS:
        messages.info(request, 'This exam has already been submitted')
        return redirect('exam_submitted', attempt_id=attempt_id)
    
    # Get exam with questions
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('student_exam_list')
    
    # Get existing answers
    existing_answers = answer_service.get_attempt_answers(attempt_id)
    answer_dict = {answer.question_id: answer for answer in existing_answers}
    
    # Calculate remaining time
    elapsed_time = (timezone.now() - attempt.started_at).total_seconds()
    total_time = exam.duration_minutes * 60
    remaining_time = max(0, total_time - elapsed_time)
    
    # Prepare questions with answers
    questions_data = []
    for question in exam.questions.all():
        question_info = {
            'question': question,
            'answer': answer_dict.get(question.id)
        }
        questions_data.append(question_info)
    
    # Count answered questions
    answered_count = sum(1 for data in questions_data if data['answer'])
    
    # Build breadcrumbs
    breadcrumbs = build_breadcrumbs(
        ('Available Exams', reverse('student_exam_list')),
        ('Review Exam', None)
    )
    
    context = {
        'attempt': attempt,
        'exam': exam,
        'questions_data': questions_data,
        'remaining_time': int(remaining_time),
        'duration_minutes': exam.duration_minutes,
        'total_questions': len(questions_data),
        'answered_count': answered_count,
        'page_breadcrumbs': breadcrumbs
    }
    return render(request, 'attempts/exam_review.html', context)


@require_http_methods(["POST"])
def save_answer_view(request, attempt_id):
    """
    Save a student's answer to a question (AJAX endpoint).
    Implements auto-save functionality with answer preservation.
    Requirements: 11.1, 11.2, 19.1
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Get attempt
    attempt = attempt_service.get_attempt(attempt_id)
    if not attempt:
        return JsonResponse({'error': 'Attempt not found'}, status=404)
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Check if attempt is still in progress (answer immutability)
    if attempt.status != AttemptStatus.IN_PROGRESS:
        return JsonResponse({'error': 'Cannot modify answers after submission'}, status=400)
    
    # Parse request data
    try:
        data = json.loads(request.body)
        question_id = data.get('question_id')
        answer_text = data.get('answer_text')
        
        if not question_id:
            return JsonResponse({'error': 'Question ID is required'}, status=400)
        
        # Save answer
        answer = answer_service.save_answer(attempt_id, question_id, answer_text)
        
        if answer:
            return JsonResponse({
                'success': True,
                'message': 'Answer saved successfully'
            })
        else:
            return JsonResponse({'error': 'Failed to save answer'}, status=500)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def submit_exam_view(request, attempt_id):
    """
    Submit an exam attempt (AJAX endpoint).
    Marks the attempt as submitted and prevents further modifications.
    Automatically grades objective questions (MCQ, True/False, Identification, Enumeration).
    Implements idempotent submission handling with proper locking.
    Supports auto-submission due to tab violations with automatic flagging.
    Requirements: 11.3, 11.4, 11.5, 19.5, 1.4, 4.4
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Get attempt
    attempt = attempt_service.get_attempt(attempt_id)
    if not attempt:
        return JsonResponse({'error': 'Attempt not found'}, status=404)
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Check if attempt is already submitted (idempotent submission)
    if attempt.status != AttemptStatus.IN_PROGRESS:
        return JsonResponse({
            'success': True,
            'message': 'Exam already submitted',
            'already_submitted': True,
            'redirect_url': f'/attempts/student/attempts/{attempt_id}/submitted/'
        })
    
    # Parse request data to check for auto_submit parameter
    auto_submit = False
    try:
        if request.body:
            data = json.loads(request.body)
            auto_submit = data.get('auto_submit', False)
    except (json.JSONDecodeError, ValueError):
        # If parsing fails, treat as normal submission
        auto_submit = False
    
    # Submit attempt with proper locking for concurrent submission handling
    # Pass auto_submit parameter to handle flagging internally
    submitted_attempt = attempt_service.submit_attempt(attempt_id, auto_submit=auto_submit)
    
    if submitted_attempt:
        # Log auto-submission if applicable
        if auto_submit:
            logger.info(f"Attempt {attempt_id} auto-submitted and flagged due to tab violations")
        
        # Auto-grade objective questions after submission
        from services.auto_grader_service import AutoGraderService
        auto_grader = AutoGraderService()
        
        try:
            # Grade the attempt automatically
            grading_success = auto_grader.grade_attempt(attempt_id)
            
            if grading_success:
                logger.info(f"Auto-grading completed for attempt {attempt_id}")
            else:
                logger.warning(f"Auto-grading failed for attempt {attempt_id}")
        except Exception as e:
            # Log error but don't fail the submission
            logger.error(f"Error during auto-grading: {e}")
        
        # Invalidate dashboard cache for this student (Requirements 11.1, 11.5)
        from services.dashboard_service import DashboardService
        DashboardService.invalidate_student_cache(student_id)
        
        return JsonResponse({
            'success': True,
            'message': 'Exam submitted and graded successfully',
            'already_submitted': False,
            'redirect_url': f'/attempts/student/attempts/{attempt_id}/submitted/'
        })
    else:
        return JsonResponse({'error': 'Failed to submit exam'}, status=500)


def exam_submitted_view(request, attempt_id):
    """
    Display submission confirmation page.
    Shows exam details and confirmation message.
    Requirements: 11.4
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        messages.error(request, 'Please log in to view this page')
        return redirect('student_login')
    
    student_id = request.session['student_id']
    
    # Get attempt with full data
    attempt = attempt_service.get_attempt_with_full_data(attempt_id)
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('student_exam_list')
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        messages.error(request, 'You do not have permission to access this page')
        return redirect('student_exam_list')
    
    # Verify attempt is submitted
    if attempt.status == AttemptStatus.IN_PROGRESS:
        messages.warning(request, 'This exam has not been submitted yet')
        return redirect('exam_take', attempt_id=attempt_id)
    
    # Get exam details
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('student_exam_list')
    
    # Get answers count
    answers = answer_service.get_attempt_answers(attempt_id)
    answers_count = len(answers)
    total_questions = exam.questions.count()
    
    # Check if results are available (for showing results link)
    show_results_link = attempt.status == AttemptStatus.GRADED
    
    # Calculate total possible points and percentage
    total_possible_points = sum(float(q.points) for q in exam.questions.all())
    percentage_score = 0
    if total_possible_points > 0:
        percentage_score = (float(attempt.total_score) / total_possible_points) * 100
    
    # Count auto-graded questions
    from exams.models import QuestionType
    auto_graded_count = sum(1 for answer in answers if answer.is_correct is not None)
    essay_count = sum(1 for answer in answers if answer.question.question_type == QuestionType.ESSAY)
    
    context = {
        'attempt': attempt,
        'exam': exam,
        'answers_count': answers_count,
        'total_questions': total_questions,
        'show_results_link': show_results_link,
        'total_score': float(attempt.total_score),
        'total_possible_points': total_possible_points,
        'percentage_score': round(percentage_score, 2),
        'auto_graded_count': auto_graded_count,
        'essay_count': essay_count
    }
    
    return render(request, 'attempts/exam_submitted.html', context)


def teacher_grading_list_view(request):
    """Show the teacher a list of students with submission summaries."""
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access grading')
        return redirect('teacher_login')

    attempts = _get_teacher_submission_attempts()
    student_rows = _build_student_rows(attempts)

    paginator = Paginator(student_rows, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    breadcrumbs = build_breadcrumbs(
        ('Dashboard', reverse('teacher_dashboard')),
        'Student Submissions'
    )

    context = {
        'page_obj': page_obj,
        'student_rows': page_obj.object_list,
        'total_students': paginator.count,
        'total_attempts': sum(row['attempt_count'] for row in student_rows),
        'total_pending': sum(row['pending_attempt_count'] for row in student_rows),
        'page_breadcrumbs': breadcrumbs,
    }

    return render(request, 'attempts/grading_list.html', context)


def teacher_student_detail_view(request, student_id):
    """Show all submissions for one student, grouped by exam."""
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access student submissions')
        return redirect('teacher_login')

    from users.models import Student

    try:
        student = Student.objects.get(pk=student_id)
    except Student.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('teacher_grading_list')

    attempts = list(_get_teacher_submission_attempts().filter(student_id=student_id))
    exam_groups = _build_exam_groups_for_student(attempts)

    total_attempts = len(attempts)
    pending_attempts = sum(1 for attempt in attempts if _attempt_pending_essay_count(attempt) > 0)
    graded_attempts = sum(1 for attempt in attempts if attempt.status == AttemptStatus.GRADED)
    average_score = 0
    if attempts:
        average_score = sum(float(attempt.total_score) for attempt in attempts) / len(attempts)

    latest_submission = None
    if attempts:
        latest_submission = max(
            (_attempt_timestamp(attempt) for attempt in attempts if _attempt_timestamp(attempt)),
            default=None,
        )

    breadcrumbs = build_breadcrumbs(
        ('Dashboard', reverse('teacher_dashboard')),
        ('Student Submissions', reverse('teacher_grading_list')),
        student.get_full_name(),
    )

    context = {
        'student': student,
        'exam_groups': exam_groups,
        'total_attempts': total_attempts,
        'pending_attempts': pending_attempts,
        'graded_attempts': graded_attempts,
        'average_score': round(average_score, 2),
        'latest_submission': latest_submission,
        'page_breadcrumbs': breadcrumbs,
    }

    return render(request, 'attempts/student_submissions_detail.html', context)


def teacher_pending_grading_view(request):
    """Show attempts that still have ungraded essays."""
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access grading')
        return redirect('teacher_login')

    pending_rows = _build_pending_grading_rows(list(_get_teacher_submission_attempts()))

    paginator = Paginator(pending_rows, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    breadcrumbs = build_breadcrumbs(
        ('Dashboard', reverse('teacher_dashboard')),
        'Pending Essay Grading'
    )

    context = {
        'page_obj': page_obj,
        'pending_rows': page_obj.object_list,
        'total_pending_attempts': paginator.count,
        'page_breadcrumbs': breadcrumbs,
    }

    return render(request, 'attempts/grading_pending.html', context)


def teacher_grading_view(request, attempt_id):
    """
    Display grading interface for essay questions.
    Shows student responses, question text, and allows score assignment.
    Requirements: 13.1, 17.1, 17.2
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access grading')
        return redirect('teacher_login')
    
    # Import grading service
    from services.grading_service import GradingService
    from users.models import Student
    from exams.models import QuestionType
    
    grading_service_instance = GradingService()
    
    # Get attempt with answers and questions
    from repositories.attempt_repository import AttemptRepository
    attempt_repository = AttemptRepository()
    attempt = attempt_repository.get_with_answers_and_questions(attempt_id)
    
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('teacher_grading_list')
    
    # Verify attempt is submitted
    if attempt.status == AttemptStatus.IN_PROGRESS:
        messages.warning(request, 'This exam has not been submitted yet')
        return redirect('teacher_grading_list')
    
    # Get student info
    try:
        student = Student.objects.get(id=attempt.student_id)
    except Student.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('teacher_grading_list')
    
    # Get exam info
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('teacher_grading_list')
    
    # Get grading status
    grading_status = grading_service_instance.get_grading_status(attempt_id)

    breadcrumbs = build_breadcrumbs(
        ('Dashboard', reverse('teacher_dashboard')),
        ('Student Submissions', reverse('teacher_grading_list')),
        (student.get_full_name(), reverse('teacher_student_detail', args=[student.id])),
        'Grade Essay',
    )
    
    # Prepare questions with answers
    questions_data = []
    for answer in attempt.answers.all():
        question = answer.question
        
        # Include all questions for context, but highlight essays
        question_info = {
            'answer': answer,
            'question': question,
            'is_essay': question.question_type == QuestionType.ESSAY,
            'is_graded': answer.is_correct is not None,
            'answer_text': answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
        }
        questions_data.append(question_info)
    
    context = {
        'attempt': attempt,
        'student': student,
        'exam': exam,
        'questions_data': questions_data,
        'grading_status': grading_status,
        'page_breadcrumbs': breadcrumbs,
    }
    
    return render(request, 'attempts/grading.html', context)


@require_http_methods(["POST"])
def grade_essay_view(request, answer_id):
    """
    Grade a single essay question (AJAX endpoint).
    Accepts score and feedback, updates answer and recalculates total score.
    Requirements: 13.2, 13.3, 13.4
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    # Import grading service
    from services.grading_service import GradingService
    
    grading_service_instance = GradingService()
    
    # Parse request data
    try:
        data = json.loads(request.body)
        points_earned = float(data.get('points_earned', 0))
        teacher_feedback = data.get('teacher_feedback', '')
        
        # Grade the essay
        result = grading_service_instance.grade_essay(
            answer_id=answer_id,
            points_earned=points_earned,
            teacher_feedback=teacher_feedback if teacher_feedback else None
        )
        
        if result:
            # Invalidate dashboard cache for this student (Requirements 11.1, 11.5)
            from services.dashboard_service import DashboardService
            DashboardService.invalidate_student_cache(result['attempt'].student_id)
            
            return JsonResponse({
                'success': True,
                'message': 'Essay graded successfully',
                'points_earned': float(result['answer'].points_earned),
                'total_score': float(result['attempt'].total_score),
                'attempt_status': result['attempt'].status
            })
        else:
            return JsonResponse({'error': 'Failed to grade essay'}, status=500)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid points value: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
def update_essay_score_view(request, answer_id):
    """
    Update the score for a previously graded essay (AJAX endpoint).
    Allows teachers to modify essay scores after initial grading.
    Requirements: 13.5
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    # Import grading service
    from services.grading_service import GradingService
    
    grading_service_instance = GradingService()
    
    # Parse request data
    try:
        data = json.loads(request.body)
        new_points_earned = float(data.get('points_earned', 0))
        new_feedback = data.get('teacher_feedback')
        
        # Modify the essay score
        result = grading_service_instance.modify_essay_score(
            answer_id=answer_id,
            new_points_earned=new_points_earned,
            new_feedback=new_feedback
        )
        
        if result:
            return JsonResponse({
                'success': True,
                'message': 'Essay score updated successfully',
                'points_earned': float(result['answer'].points_earned),
                'total_score': float(result['attempt'].total_score),
                'attempt_status': result['attempt'].status
            })
        else:
            return JsonResponse({'error': 'Failed to update essay score'}, status=500)
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid points value: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



def student_results_view(request, attempt_id):
    """
    Display exam results for students.
    Shows scores, questions, submitted answers, correct answers, and teacher feedback.
    Only displays results for completed and graded exams.
    Requirements: 15.1, 15.2, 15.3, 15.4, 15.5
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        messages.error(request, 'Please log in to view results')
        return redirect('student_login')
    
    student_id = request.session['student_id']
    
    # Get attempt with full data
    attempt = attempt_service.get_attempt_with_full_data(attempt_id)
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('student_exam_list')
    
    # Verify this attempt belongs to the logged-in student
    if attempt.student_id != student_id:
        messages.error(request, 'You do not have permission to access these results')
        return redirect('student_exam_list')
    
    # Hide results for ungraded exams (Requirements 15.1, 15.4)
    if attempt.status != AttemptStatus.GRADED:
        messages.info(request, 'Results are not yet available. Your exam is still being graded.')
        return redirect('student_exam_list')
    
    # Get exam with questions
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('student_exam_list')
    
    # Get answers
    answers = answer_service.get_attempt_answers(attempt_id)
    answer_dict = {answer.question_id: answer for answer in answers}
    
    # Calculate total possible points
    total_possible_points = sum(float(q.points) for q in exam.questions.all())
    
    # Calculate percentage score (Requirements 15.5)
    percentage_score = 0
    if total_possible_points > 0:
        percentage_score = (float(attempt.total_score) / total_possible_points) * 100
    
    # Prepare questions with answers and correct answers (Requirements 15.2, 15.3)
    from exams.models import QuestionType
    questions_data = []
    for question in exam.questions.all():
        answer = answer_dict.get(question.id)
        
        # Format student answer based on question type
        student_answer_display = None
        if answer:
            if question.question_type == QuestionType.MCQ:
                # Get the selected option text
                answer_key = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
                student_answer_display = answer_key
                # Find the option text
                for option in question.options:
                    if option.get('key') == answer_key:
                        student_answer_display = f"{answer_key}. {option.get('value')}"
                        break
            elif question.question_type == QuestionType.TRUE_FALSE:
                answer_value = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
                student_answer_display = str(answer_value).title()
            elif question.question_type == QuestionType.ENUMERATION:
                answer_value = answer.answer_text.get('value', []) if isinstance(answer.answer_text, dict) else answer.answer_text
                if isinstance(answer_value, list):
                    student_answer_display = ', '.join(answer_value)
                else:
                    student_answer_display = str(answer_value)
            else:
                # IDENTIFICATION or ESSAY
                student_answer_display = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
        
        # Format correct answer based on question type
        correct_answer_display = None
        if question.question_type == QuestionType.MCQ:
            # Show the correct option
            correct_key = question.correct_answer
            for option in question.options:
                if option.get('key') == correct_key:
                    correct_answer_display = f"{correct_key}. {option.get('value')}"
                    break
            if not correct_answer_display:
                correct_answer_display = correct_key
        elif question.question_type == QuestionType.TRUE_FALSE:
            correct_answer_display = str(question.correct_answer).title()
        elif question.question_type == QuestionType.ENUMERATION:
            if isinstance(question.correct_answer, list):
                correct_answer_display = ', '.join(question.correct_answer)
            else:
                correct_answer_display = str(question.correct_answer)
        elif question.question_type == QuestionType.IDENTIFICATION:
            if isinstance(question.correct_answer, list):
                correct_answer_display = ' / '.join(question.correct_answer)
            else:
                correct_answer_display = str(question.correct_answer)
        elif question.question_type == QuestionType.ESSAY:
            # Essays don't have a single correct answer
            correct_answer_display = None
        
        question_info = {
            'question': question,
            'answer': answer,
            'student_answer_display': student_answer_display,
            'correct_answer_display': correct_answer_display,
            'is_correct': answer.is_correct if answer else None,
            'points_earned': float(answer.points_earned) if answer else 0,
            'teacher_feedback': answer.teacher_feedback if answer and answer.teacher_feedback else None,
            'is_essay': question.question_type == QuestionType.ESSAY
        }
        questions_data.append(question_info)
    
    context = {
        'attempt': attempt,
        'exam': exam,
        'questions_data': questions_data,
        'total_score': float(attempt.total_score),
        'total_possible_points': total_possible_points,
        'percentage_score': round(percentage_score, 2),
        'submitted_at': attempt.submitted_at
    }
    
    return render(request, 'attempts/student_results.html', context)


def teacher_dashboard_view(request):
    """
    Display teacher dashboard with all student attempts and analytics.
    Shows all attempts with filtering by exam, student, and class.
    Displays class statistics (average, highest, lowest scores).
    Requirements: 14.1, 14.2, 14.3, 14.4, 14.5, 5.1, 5.2, 5.3, 3.4
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access the dashboard')
        return redirect('teacher_login')
    
    from repositories.attempt_repository import AttemptRepository
    from users.models import Student, Class
    from django.db.models import Avg, Max, Min, Count
    from services.dashboard_service import DashboardService
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    attempt_repository = AttemptRepository()
    dashboard_service = DashboardService()
    
    # Get filter parameters
    exam_filter = request.GET.get('exam', '')
    student_filter = request.GET.get('student', '')
    status_filter = request.GET.get('status', '')
    class_filter = request.GET.get('class', '')  # New class filter (Requirement 5.1)
    
    # Get all submitted and graded attempts with optimized queries (Requirement 9.5)
    # Use select_related to fetch related student and exam in a single query
    # Use prefetch_related to efficiently fetch exam questions
    attempts = Attempt.objects.filter(
        status__in=[AttemptStatus.SUBMITTED, AttemptStatus.GRADED]
    ).select_related('student', 'student__class_assigned', 'exam').prefetch_related('exam__questions').order_by('-submitted_at')
    
    # Apply filters
    if exam_filter:
        try:
            exam_id = int(exam_filter)
            attempts = attempts.filter(exam_id=exam_id)
        except ValueError:
            pass
    
    if student_filter:
        try:
            student_id = int(student_filter)
            attempts = attempts.filter(student_id=student_id)
        except ValueError:
            pass
    
    if status_filter:
        attempts = attempts.filter(status=status_filter)
    
    # Apply class filter (Requirement 5.1, 5.2)
    if class_filter:
        try:
            class_id = int(class_filter)
            attempts = attempts.filter(student__class_assigned_id=class_id)
        except ValueError:
            pass
    
    # Prepare attempt data
    attempts_data = []
    for attempt in attempts:
        student = attempt.student
        exam = attempt.exam
        
        # Calculate percentage
        total_possible = sum(float(q.points) for q in exam.questions.all())
        percentage = 0
        if total_possible > 0:
            percentage = (float(attempt.total_score) / total_possible) * 100
        
        attempts_data.append({
            'attempt': attempt,
            'student': student,
            'exam': exam,
            'percentage': round(percentage, 2),
            'total_possible': total_possible,
            'class': student.class_assigned  # Include class info
        })
    
    # Implement pagination (Requirement 3.4)
    page_number = request.GET.get('page', 1)
    items_per_page = 10  # Show 10 attempts per page
    paginator = Paginator(attempts_data, items_per_page)
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    # Use paginated data for display
    paginated_attempts_data = page_obj.object_list
    
    # Calculate class statistics (Requirements 14.5, 5.2, 5.3)
    statistics = {}
    if attempts.exists():
        stats = attempts.aggregate(
            average=Avg('total_score'),
            highest=Max('total_score'),
            lowest=Min('total_score'),
            total_attempts=Count('id')
        )
        statistics = {
            'average': round(float(stats['average'] or 0), 2),
            'highest': float(stats['highest'] or 0),
            'lowest': float(stats['lowest'] or 0),
            'total_attempts': stats['total_attempts']
        }
    
    # Get class-grouped statistics (Requirement 5.3)
    class_statistics = {}
    if class_filter:
        # If filtering by class, get statistics for that class
        try:
            class_id = int(class_filter)
            class_stats_result = dashboard_service.get_statistics_by_class(class_id)
            if class_stats_result.is_success():
                class_statistics = {class_id: class_stats_result.value}
        except ValueError:
            pass
    else:
        # Get statistics for all classes
        all_classes = Class.objects.all()
        for cls in all_classes:
            class_stats_result = dashboard_service.get_statistics_by_class(cls.id)
            if class_stats_result.is_success():
                class_statistics[cls.id] = class_stats_result.value
    
    # Get all exams, students, and classes for filter dropdowns
    exams = exam_service.get_all_exams()
    students = Student.objects.all().order_by('last_name', 'first_name')
    classes = Class.objects.all().order_by('grade_level', 'strand', 'section')  # New class dropdown (Requirement 5.1)
    
    # Calculate chart data for visualizations
    # 1. Total counts
    total_students = students.count()
    total_exams = len(exams) if isinstance(exams, list) else exams.count()
    
    # 2. Pass/Fail by exam (passing threshold: 60%)
    exam_performance = {}
    for exam in exams:
        exam_attempts = [data for data in attempts_data if data['exam'].id == exam.id]
        if exam_attempts:
            passers = sum(1 for data in exam_attempts if data['percentage'] >= 60)
            failers = len(exam_attempts) - passers
            exam_performance[exam.title] = {
                'passers': passers,
                'failers': failers,
                'total': len(exam_attempts)
            }
    
    # 3. Overall pass/fail statistics
    total_passers = sum(1 for data in attempts_data if data['percentage'] >= 60)
    total_failers = len(attempts_data) - total_passers
    
    # Get passing rate by subject per section
    try:
        # Check if teacher profile exists
        if hasattr(request.user, 'teacher_profile'):
            teacher = request.user.teacher_profile
            # Pass the teacher's primary key (not user_id)
            passing_rate_result = dashboard_service.get_passing_rate_by_subject_per_section(teacher.pk)
            if passing_rate_result.is_success():
                passing_rate_data = passing_rate_result.value
            else:
                passing_rate_data = {'sections': [], 'subjects': [], 'data': {}}
        else:
            passing_rate_data = {'sections': [], 'subjects': [], 'data': {}}
    except Exception as e:
        logger.error(f"Error getting passing rate data: {str(e)}")
        passing_rate_data = {'sections': [], 'subjects': [], 'data': {}}
    
    # Ensure passing_rate_data is JSON serializable
    passing_rate_json = passing_rate_data

    class_statistics_json = {
        str(class_id): {
            'class_name': stats.get('class_name', ''),
            'average': stats.get('average', 0),
            'student_count': stats.get('student_count', 0),
            'total_attempts': stats.get('total_attempts', 0)
        }
        for class_id, stats in class_statistics.items()
    }
    
    context = {
        'attempts_data': paginated_attempts_data,  # Use paginated data
        'page_obj': page_obj,  # Add pagination object (Requirement 3.4)
        'statistics': statistics,
        'exams': exams,
        'students': students,
        'classes': classes,  # New class list for filter dropdown (Requirement 5.1)
        'exam_filter': exam_filter,
        'student_filter': student_filter,
        'status_filter': status_filter,
        'class_filter': class_filter,  # New class filter value (Requirement 5.1)
        'total_attempts': len(attempts_data),  # Keep total count for display
        # Chart data
        'total_students': total_students,
        'total_exams': total_exams,
        'exam_performance': exam_performance,
        'exam_performance_json': exam_performance,
        'total_passers': total_passers,
        'total_failers': total_failers,
        # Class statistics (Requirements 5.2, 5.3)
        'class_statistics': class_statistics,
        'class_statistics_json': class_statistics_json,
        # Passing rate by subject per section
        'passing_rate_data': passing_rate_data,
        'passing_rate_json': passing_rate_json,
        # Dashboard configuration for JavaScript
        'dashboard_config': {
            'total_students': total_students,
            'total_exams': total_exams,
            'total_passers': total_passers,
            'total_failers': total_failers
        }
    }
    
    return render(request, 'attempts/teacher_dashboard.html', context)


def teacher_attempt_detail_view(request, attempt_id):
    """
    Display detailed view of a specific attempt for teachers.
    Shows all questions, answers, scores, and grading information.
    Requirements: 14.4
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access attempt details')
        return redirect('teacher_login')
    
    from repositories.attempt_repository import AttemptRepository
    from exams.models import QuestionType
    
    attempt_repository = AttemptRepository()
    
    # Get attempt with full data
    attempt = attempt_repository.get_with_answers_and_questions(attempt_id)
    if not attempt:
        messages.error(request, 'Attempt not found')
        return redirect('teacher_dashboard')
    
    # Get student and exam info
    student = attempt.student
    exam = exam_service.get_exam_with_questions(attempt.exam_id)
    if not exam:
        messages.error(request, 'Exam not found')
        return redirect('teacher_dashboard')
    
    # Get answers
    answers = answer_service.get_attempt_answers(attempt_id)
    answer_dict = {answer.question_id: answer for answer in answers}
    
    # Calculate total possible points
    total_possible_points = sum(float(q.points) for q in exam.questions.all())
    
    # Calculate percentage score
    percentage_score = 0
    if total_possible_points > 0:
        percentage_score = (float(attempt.total_score) / total_possible_points) * 100
    
    # Prepare questions with answers
    questions_data = []
    for question in exam.questions.all():
        answer = answer_dict.get(question.id)
        
        # Format student answer based on question type
        student_answer_display = None
        if answer:
            if question.question_type == QuestionType.MCQ:
                answer_key = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
                student_answer_display = answer_key
                for option in question.options:
                    if option.get('key') == answer_key:
                        student_answer_display = f"{answer_key}. {option.get('value')}"
                        break
            elif question.question_type == QuestionType.TRUE_FALSE:
                answer_value = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
                student_answer_display = str(answer_value).title()
            elif question.question_type == QuestionType.ENUMERATION:
                answer_value = answer.answer_text.get('value', []) if isinstance(answer.answer_text, dict) else answer.answer_text
                if isinstance(answer_value, list):
                    student_answer_display = ', '.join(answer_value)
                else:
                    student_answer_display = str(answer_value)
            else:
                student_answer_display = answer.answer_text.get('value', '') if isinstance(answer.answer_text, dict) else answer.answer_text
        
        # Format correct answer
        correct_answer_display = None
        if question.question_type == QuestionType.MCQ:
            correct_key = question.correct_answer
            for option in question.options:
                if option.get('key') == correct_key:
                    correct_answer_display = f"{correct_key}. {option.get('value')}"
                    break
            if not correct_answer_display:
                correct_answer_display = correct_key
        elif question.question_type == QuestionType.TRUE_FALSE:
            correct_answer_display = str(question.correct_answer).title()
        elif question.question_type == QuestionType.ENUMERATION:
            if isinstance(question.correct_answer, list):
                correct_answer_display = ', '.join(question.correct_answer)
            else:
                correct_answer_display = str(question.correct_answer)
        elif question.question_type == QuestionType.IDENTIFICATION:
            if isinstance(question.correct_answer, list):
                correct_answer_display = ' / '.join(question.correct_answer)
            else:
                correct_answer_display = str(question.correct_answer)
        elif question.question_type == QuestionType.ESSAY:
            correct_answer_display = None
        
        question_info = {
            'question': question,
            'answer': answer,
            'student_answer_display': student_answer_display,
            'correct_answer_display': correct_answer_display,
            'is_correct': answer.is_correct if answer else None,
            'points_earned': float(answer.points_earned) if answer else 0,
            'teacher_feedback': answer.teacher_feedback if answer and answer.teacher_feedback else None,
            'is_essay': question.question_type == QuestionType.ESSAY
        }
        questions_data.append(question_info)
    
    context = {
        'attempt': attempt,
        'student': student,
        'exam': exam,
        'questions_data': questions_data,
        'total_score': float(attempt.total_score),
        'total_possible_points': total_possible_points,
        'percentage_score': round(percentage_score, 2),
        'submitted_at': attempt.submitted_at
    }
    
    return render(request, 'attempts/attempt_detail.html', context)


@require_http_methods(["GET"])
def recover_interrupted_attempt_view(request, exam_id):
    """
    Recover an interrupted exam attempt for a student (AJAX endpoint).
    Restores the attempt state and all saved answers after connection drop.
    Requirements: 19.1, 19.2
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Attempt to recover interrupted attempt
    recovery_data = data_integrity_service.recover_interrupted_attempt(
        student_id, exam_id
    )
    
    if recovery_data:
        return JsonResponse({
            'success': True,
            'recovered': True,
            'attempt_id': recovery_data['attempt_id'],
            'started_at': recovery_data['started_at'].isoformat(),
            'answers': recovery_data['answers'],
            'message': 'Exam state recovered successfully'
        })
    else:
        return JsonResponse({
            'success': True,
            'recovered': False,
            'message': 'No interrupted attempt found'
        })



@require_http_methods(["POST"])
def record_tab_switch_view(request, attempt_id):
    """
    Record a tab switch violation (AJAX endpoint).
    Tracks when student switches away from exam tab and issues warnings.
    Returns current warning count and auto-submit flag.
    Requirements: 2.1, 3.1, 3.4, 8.1, 8.2, 8.5
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        logger.warning(f"Unauthenticated attempt to record tab switch for attempt {attempt_id}")
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Validate attempt_id is a valid integer
    try:
        attempt_id = int(attempt_id)
    except (ValueError, TypeError):
        logger.error(f"Invalid attempt_id format: {attempt_id}")
        return JsonResponse({'error': 'Invalid attempt ID'}, status=400)
    
    # Get attempt with error handling
    try:
        attempt = attempt_service.get_attempt(attempt_id)
        if not attempt:
            logger.error(f"Attempt {attempt_id} not found")
            return JsonResponse({'error': 'Attempt not found'}, status=404)
    except Exception as e:
        logger.error(f"Database error retrieving attempt {attempt_id}: {e}")
        return JsonResponse({'error': 'Database error'}, status=500)
    
    # Verify this attempt belongs to the logged-in student (attempt ownership verification)
    if attempt.student_id != student_id:
        logger.warning(
            f"Permission denied: Student {student_id} attempted to record violation "
            f"for attempt {attempt_id} belonging to student {attempt.student_id}"
        )
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Check if attempt is still in progress
    if attempt.status != AttemptStatus.IN_PROGRESS:
        logger.warning(
            f"Cannot record violation for attempt {attempt_id}: "
            f"Status is {attempt.status}, not IN_PROGRESS"
        )
        return JsonResponse({'error': 'Cannot record violations for submitted exams'}, status=400)
    
    # Parse request data with comprehensive error handling
    try:
        # Validate request body exists
        if not request.body:
            logger.error(f"Empty request body for attempt {attempt_id}")
            return JsonResponse({'error': 'Request body is required'}, status=400)
        
        data = json.loads(request.body)
        
        # Validate data is a dictionary
        if not isinstance(data, dict):
            logger.error(f"Invalid data type for attempt {attempt_id}: {type(data)}")
            return JsonResponse({'error': 'Invalid data format'}, status=400)
        
        violated_at_str = data.get('violated_at')
        
        # Parse timestamp if provided, otherwise use current time
        violated_at = None
        if violated_at_str:
            try:
                from django.utils.dateparse import parse_datetime
                violated_at = parse_datetime(violated_at_str)
                if not violated_at:
                    # Try parsing as ISO format
                    from datetime import datetime
                    violated_at = datetime.fromisoformat(violated_at_str.replace('Z', '+00:00'))
                
                # Validate timestamp is not in the future
                if violated_at and violated_at > timezone.now():
                    logger.warning(f"Future timestamp provided for attempt {attempt_id}: {violated_at_str}")
                    violated_at = None  # Use current time instead
            except (ValueError, TypeError) as e:
                logger.warning(f"Invalid timestamp format for attempt {attempt_id}: {violated_at_str}, error: {e}")
                violated_at = None  # Use current time
        
        # Record the tab switch violation with transaction for concurrent handling
        from django.db import transaction
        
        try:
            with transaction.atomic():
                result = tab_monitoring_service.record_tab_switch(
                    attempt_id=attempt_id,
                    violated_at=violated_at
                )
                
                # If this is the 4th violation, flag the attempt
                if result['should_auto_submit']:
                    flag_success = tab_monitoring_service.flag_attempt_for_cheating(
                        attempt_id=attempt_id,
                        reason="Auto-submitted after 4 tab switches"
                    )
                    if not flag_success:
                        logger.error(f"Failed to flag attempt {attempt_id} after 4th violation")
        except Exception as e:
            logger.error(f"Transaction error recording tab switch for attempt {attempt_id}: {e}")
            return JsonResponse({'error': 'Failed to record violation'}, status=500)
        
        logger.info(
            f"Tab switch recorded for attempt {attempt_id} by student {student_id}: "
            f"Warning {result['warning_number']}/{result['total_warnings']}"
        )
        
        return JsonResponse({
            'success': True,
            'warning_number': result['warning_number'],
            'total_warnings': result['total_warnings'],
            'should_auto_submit': result['should_auto_submit'],
            'violation_id': result['violation_id']
        })
    
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error for attempt {attempt_id}: {e}")
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(
            f"Unexpected error recording tab switch for attempt {attempt_id}: {e}",
            exc_info=True
        )
        return JsonResponse({'error': 'Internal server error'}, status=500)


@require_http_methods(["GET"])
def get_tab_violations_view(request, attempt_id):
    """
    Get current violation count for an attempt (AJAX endpoint).
    Used for state restoration after page refresh.
    Returns violation count and flagged status.
    Requirements: 3.4, 8.1, 8.2, 8.5
    """
    # Check if user is authenticated as student
    if not hasattr(request, 'session') or 'student_id' not in request.session:
        logger.warning(f"Unauthenticated attempt to get violations for attempt {attempt_id}")
        return JsonResponse({'error': 'Not authenticated'}, status=401)
    
    student_id = request.session['student_id']
    
    # Validate attempt_id is a valid integer
    try:
        attempt_id = int(attempt_id)
    except (ValueError, TypeError):
        logger.error(f"Invalid attempt_id format: {attempt_id}")
        return JsonResponse({'error': 'Invalid attempt ID'}, status=400)
    
    # Get attempt with error handling
    try:
        attempt = attempt_service.get_attempt(attempt_id)
        if not attempt:
            logger.error(f"Attempt {attempt_id} not found")
            return JsonResponse({'error': 'Attempt not found'}, status=404)
    except Exception as e:
        logger.error(f"Database error retrieving attempt {attempt_id}: {e}")
        return JsonResponse({'error': 'Database error'}, status=500)
    
    # Verify this attempt belongs to the logged-in student (attempt ownership verification)
    if attempt.student_id != student_id:
        logger.warning(
            f"Permission denied: Student {student_id} attempted to get violations "
            f"for attempt {attempt_id} belonging to student {attempt.student_id}"
        )
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Get violation count with error handling
        violation_count = tab_monitoring_service.get_violation_count(attempt_id)
        
        # Validate violation count is a valid number
        if not isinstance(violation_count, int) or violation_count < 0:
            logger.error(f"Invalid violation count for attempt {attempt_id}: {violation_count}")
            violation_count = 0
        
        logger.info(
            f"Retrieved violation count for attempt {attempt_id} by student {student_id}: {violation_count}"
        )
        
        return JsonResponse({
            'success': True,
            'violation_count': violation_count,
            'is_flagged': bool(attempt.is_flagged),
            'flag_reason': attempt.flag_reason if attempt.is_flagged else None
        })
    
    except Exception as e:
        logger.error(
            f"Unexpected error getting violations for attempt {attempt_id}: {e}",
            exc_info=True
        )
        return JsonResponse({'error': 'Internal server error'}, status=500)



def view_activity_log_view(request, attempt_id):
    """
    Display detailed activity log for a specific attempt.
    Shows timeline of tab violations, warnings, and summary statistics.
    Accessible from grading list, student history, and class results.
    Requirements: 5.1, 5.2, 5.3, 5.4, 5.5, 6.4, 8.1, 8.2
    """
    # Check if user is authenticated as teacher
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher to access activity logs')
        return redirect('teacher_login')
    
    # Validate attempt_id
    try:
        attempt_id = int(attempt_id)
    except (ValueError, TypeError):
        logger.error(f"Invalid attempt_id format in activity log view: {attempt_id}")
        messages.error(request, 'Invalid attempt ID')
        return redirect('teacher_grading_list')
    
    # Get formatted activity log from service with error handling
    try:
        activity_data = activity_log_service.get_formatted_activity_log(attempt_id)
    except Exception as e:
        logger.error(
            f"Error retrieving activity log for attempt {attempt_id}: {e}",
            exc_info=True
        )
        messages.error(request, 'Failed to load activity log. Please try again.')
        return redirect('teacher_grading_list')
    
    # Check if attempt was found
    if not activity_data.get('found', False):
        error_message = activity_data.get('error', 'Attempt not found')
        logger.warning(f"Activity log not found for attempt {attempt_id}: {error_message}")
        messages.error(request, error_message)
        return redirect('teacher_grading_list')
    
    # Extract data from service response with validation
    try:
        attempt = activity_data.get('attempt')
        student = activity_data.get('student')
        exam = activity_data.get('exam')
        violations = activity_data.get('violations', [])
        summary = activity_data.get('summary', {})
        timeline = activity_data.get('timeline', [])
        
        # Validate required data
        if not attempt or not student or not exam:
            logger.error(f"Missing required data in activity log for attempt {attempt_id}")
            messages.error(request, 'Incomplete activity log data')
            return redirect('teacher_grading_list')
        
        # Build breadcrumbs
        breadcrumbs = build_breadcrumbs(
            ('Grading', 'teacher_grading_list'),
            f'Activity Log - {student.get_full_name()}'
        )
        
        context = {
            'attempt': attempt,
            'student': student,
            'exam': exam,
            'violations': violations,
            'summary': summary,
            'timeline': timeline,
            'breadcrumbs': breadcrumbs
        }
        
        return render(request, 'attempts/activity_log.html', context)
    
    except Exception as e:
        logger.error(
            f"Error processing activity log data for attempt {attempt_id}: {e}",
            exc_info=True
        )
        messages.error(request, 'Failed to display activity log. Please try again.')
        return redirect('teacher_grading_list')


@require_http_methods(["GET"])
def export_scores_excel_view(request):
    """
    Export student scores to Excel with each exam on a separate worksheet.
    Includes brand logo header on each sheet.
    """
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher')
        return redirect('teacher_login')

    import os
    from io import BytesIO
    from django.conf import settings
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    from users.models import Student
    from exams.models import Exam

    exams = Exam.objects.filter(
        attempts__status=AttemptStatus.GRADED
    ).distinct().prefetch_related('questions')

    if not exams.exists():
        messages.warning(request, 'No graded exams to export')
        return redirect('teacher_dashboard')

    wb = Workbook()
    wb.remove(wb.active)

    brand_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'brand.png')
    has_brand = os.path.isfile(brand_path)

    header_font = Font(name='Calibri', bold=True, size=12)
    title_font = Font(name='Calibri', bold=True, size=14)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_text = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for exam in exams:
        title = exam.title[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
        ws = wb.create_sheet(title=title)

        row = 1
        if has_brand:
            img = XlImage(brand_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
            row = 5

        ws.cell(row=row, column=1, value='Seamless Exam System - Score Report')
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        ws.cell(row=row, column=1, value=f'Exam: {exam.title}')
        ws.cell(row=row, column=1).font = header_font
        row += 1

        total_possible = sum(float(q.points) for q in exam.questions.all())
        ws.cell(row=row, column=1, value=f'Total Possible Points: {total_possible}')
        row += 2

        headers = ['#', 'Student Name', 'School ID', 'Class', 'Score', 'Percentage', 'Status', 'Submitted At']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        attempts = Attempt.objects.filter(
            exam=exam,
            status=AttemptStatus.GRADED
        ).select_related('student', 'student__class_assigned').order_by(
            'student__last_name', 'student__first_name'
        )

        for idx, attempt in enumerate(attempts, 1):
            student = attempt.student
            score = float(attempt.total_score)
            percentage = (score / total_possible * 100) if total_possible > 0 else 0
            class_name = student.class_assigned.name if student.class_assigned else 'N/A'

            row_data = [
                idx,
                f'{student.last_name}, {student.first_name}',
                student.school_id,
                class_name,
                score,
                round(percentage, 1),
                'Passed' if percentage >= 60 else 'Failed',
                attempt.submitted_at.strftime('%Y-%m-%d %H:%M') if attempt.submitted_at else ''
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in (5, 6):
                    cell.alignment = Alignment(horizontal='center')
                if col_idx == 7:
                    cell.alignment = Alignment(horizontal='center')
                    if value == 'Passed':
                        cell.font = Font(color='006100')
                    else:
                        cell.font = Font(color='9C0006')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value=f'Total Students: {attempts.count()}')
        ws.cell(row=row, column=1).font = Font(bold=True)

        col_widths = [5, 30, 15, 20, 10, 12, 10, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_scores_report.xlsx"'
    return response


@require_http_methods(["GET"])
def export_accounts_excel_view(request):
    """
    Export student accounts to Excel with each class on a separate worksheet.
    Includes brand logo header, class name, and student credentials.
    """
    if not request.user.is_authenticated:
        messages.error(request, 'Please log in as a teacher')
        return redirect('teacher_login')

    import os
    from io import BytesIO
    from django.conf import settings
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    from users.models import Student, Class

    classes = Class.objects.filter(teacher=request.user.teacher_profile).prefetch_related('students').order_by('grade_level', 'strand', 'section')

    if not classes.exists():
        messages.warning(request, 'No classes found to export')
        return redirect('teacher_dashboard')

    wb = Workbook()
    wb.remove(wb.active)

    brand_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'brand.png')
    has_brand = os.path.isfile(brand_path)

    header_font = Font(name='Calibri', bold=True, size=12)
    title_font = Font(name='Calibri', bold=True, size=14)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_text = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def generate_default_password(school_id, last_name):
        digits = ''.join(filter(str.isdigit, school_id))[:4]
        if len(digits) < 4:
            digits = digits.ljust(4, '0')
        letters = ''.join(filter(str.isalpha, last_name))[:4].upper()
        if len(letters) < 4:
            letters = letters.ljust(4, 'X')
        return digits + letters

    for cls in classes:
        sheet_title = f"{cls.grade_level} {cls.strand} {cls.section}"[:31]
        sheet_title = sheet_title.replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
        ws = wb.create_sheet(title=sheet_title)

        row = 1
        if has_brand:
            img = XlImage(brand_path)
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
            row = 5

        ws.cell(row=row, column=1, value='Seamless Exam System - Student Accounts')
        ws.cell(row=row, column=1).font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        ws.cell(row=row, column=1, value=f'Class: {cls.grade_level} - {cls.strand} - {cls.section}')
        ws.cell(row=row, column=1).font = header_font
        row += 2

        headers = ['#', 'First Name', 'Last Name', 'Username (School ID)', 'Default Password']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        students = cls.students.all().order_by('last_name', 'first_name')

        for idx, student in enumerate(students, 1):
            password = generate_default_password(student.school_id, student.last_name)
            row_data = [
                idx,
                student.first_name,
                student.last_name,
                student.school_id,
                password
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value=f'Total Students: {students.count()}')
        ws.cell(row=row, column=1).font = Font(bold=True)

        col_widths = [5, 20, 20, 25, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_accounts.xlsx"'
    return response
